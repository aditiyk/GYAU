import logging
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends
from typing import List, Optional
from pydantic import BaseModel
import json
import pandas as pd
import io
import datetime
import uuid

from planner.parser import parse_image_timetable, parse_csv_timetable
from planner.prioritizer import compute_priority
from planner.scheduler import generate_schedule
from planner.recommendation import generate_planner_insights
from models.planner import PlannerImport, ImportUpdateResponse
from core.security import get_current_user, UserContext
from db.database import get_db

logger = logging.getLogger(__name__)
router = APIRouter()

# Global memory for the generated schedule in demo
current_schedule = []
current_insights = []

class GenerateRequest(BaseModel):
    tasks: List[dict]

import io
import openpyxl

import logging

logger = logging.getLogger(__name__)

def is_day_label(val: str) -> bool:
    if not val:
        return False
    val_lower = str(val).lower().strip()
    return val_lower[:3] in ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun'] or val_lower == 'day name'

def is_header_label(val: str) -> bool:
    if not val:
        return False
    val_lower = str(val).lower().strip()
    if is_day_label(val):
        return True
    if 'time' in val_lower or 'day' in val_lower:
        return True
    import re
    # Check if it looks like a time label e.g., 8:00 AM, 09:00, 14:00
    if re.search(r'\d{1,2}:\d{2}', val_lower) or re.search(r'\d{1,2}\s*[ap]m', val_lower):
        return True
    return False

def parse_excel_timetable(content: bytes) -> List[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    
    # Debug logging
    logger.info("Excel Grid Dump:")
    for row in ws.iter_rows(values_only=True):
        logger.info(row)
        
    # Detect orientation
    # Let's check the first few rows and columns
    is_horizontal = False
    
    # Check first column for day names (Horizontal)
    days_in_col = 0
    for r in range(1, min(ws.max_row, 10) + 1):
        val = ws.cell(row=r, column=1).value
        if is_day_label(str(val)):
            days_in_col += 1
            
    # Check first row for day names (Vertical)
    days_in_row = 0
    for c in range(1, min(ws.max_column, 10) + 1):
        val = ws.cell(row=1, column=c).value
        if is_day_label(str(val)):
            days_in_row += 1
            
    if days_in_col >= 3 and days_in_row < 3:
        is_horizontal = True
        logger.info("Detected Horizontal layout (Days in column, Times in row)")
    else:
        logger.info("Detected Vertical layout (Days in row, Times in column)")

    # Build merged map
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = {
                    "min_row": min_row,
                    "max_row": max_row,
                    "min_col": min_col,
                    "max_col": max_col
                }

    extracted = []
    processed_merged_ranges = set()

    if not is_horizontal:
        # Vertical: Days in row, times in column 1
        days_map = {}
        header_row_idx = 1
        for r in range(1, min(ws.max_row, 5) + 1):
            found = 0
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and is_day_label(str(val)):
                    days_map[c] = str(val).strip()
                    found += 1
            if found >= 3:
                header_row_idx = r
                break
                
        times_map = {}
        for r in range(header_row_idx + 1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val:
                if hasattr(val, "strftime"):
                    times_map[r] = val.strftime("%I:%M %p")
                else:
                    times_map[r] = str(val).strip()
                    
        for r in range(header_row_idx + 1, ws.max_row + 1):
            if r not in times_map:
                continue
            start_time = times_map[r]
            for c in range(2, ws.max_column + 1):
                if c not in days_map:
                    continue
                day = days_map[c]
                
                if (r, c) in merged_map:
                    m = merged_map[(r, c)]
                    block_id = (m["min_row"], m["min_col"])
                    if block_id in processed_merged_ranges:
                        continue
                    processed_merged_ranges.add(block_id)
                    val = ws.cell(row=m["min_row"], column=m["min_col"]).value
                    if not val or not str(val).strip() or is_header_label(str(val)):
                        continue
                    duration_minutes = (m["max_row"] - m["min_row"] + 1) * 60
                    task = {
                        "title": str(val).strip(),
                        "category": "Class",
                        "fixed_or_flexible": "Fixed",
                        "completed": False,
                        "day": day,
                        "start_time": times_map.get(m["min_row"], start_time),
                        "duration_minutes": duration_minutes
                    }
                    extracted.append(task)
                    logger.info(f"Parsed task: {task}")
                else:
                    val = ws.cell(row=r, column=c).value
                    if val and str(val).strip() and not is_header_label(str(val)):
                        task = {
                            "title": str(val).strip(),
                            "category": "Class",
                            "fixed_or_flexible": "Fixed",
                            "completed": False,
                            "day": day,
                            "start_time": start_time,
                            "duration_minutes": 60
                        }
                        extracted.append(task)
                        logger.info(f"Parsed task: {task}")
    else:
        # Horizontal: Times in row, days in column 1
        times_map = {}
        header_row_idx = 1
        for r in range(1, min(ws.max_row, 5) + 1):
            found = 0
            for c in range(2, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and (is_header_label(str(val)) or hasattr(val, "strftime")):
                    if hasattr(val, "strftime"):
                        times_map[c] = val.strftime("%I:%M %p")
                    else:
                        times_map[c] = str(val).strip()
                    found += 1
            if found >= 3:
                header_row_idx = r
                break
                
        days_map = {}
        for r in range(header_row_idx + 1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val and is_day_label(str(val)):
                days_map[r] = str(val).strip()
                
        for r in range(header_row_idx + 1, ws.max_row + 1):
            if r not in days_map:
                continue
            day = days_map[r]
            for c in range(2, ws.max_column + 1):
                if c not in times_map:
                    continue
                start_time = times_map[c]
                
                if (r, c) in merged_map:
                    m = merged_map[(r, c)]
                    block_id = (m["min_row"], m["min_col"])
                    if block_id in processed_merged_ranges:
                        continue
                    processed_merged_ranges.add(block_id)
                    val = ws.cell(row=m["min_row"], column=m["min_col"]).value
                    if not val or not str(val).strip() or is_header_label(str(val)):
                        continue
                    # Duration in horizontal means it spans columns
                    duration_minutes = (m["max_col"] - m["min_col"] + 1) * 60
                    task = {
                        "title": str(val).strip(),
                        "category": "Class",
                        "fixed_or_flexible": "Fixed",
                        "completed": False,
                        "day": day,
                        "start_time": times_map.get(m["min_col"], start_time),
                        "duration_minutes": duration_minutes
                    }
                    extracted.append(task)
                    logger.info(f"Parsed task: {task}")
                else:
                    val = ws.cell(row=r, column=c).value
                    if val and str(val).strip() and not is_header_label(str(val)):
                        task = {
                            "title": str(val).strip(),
                            "category": "Class",
                            "fixed_or_flexible": "Fixed",
                            "completed": False,
                            "day": day,
                            "start_time": start_time,
                            "duration_minutes": 60
                        }
                        extracted.append(task)
                        logger.info(f"Parsed task: {task}")

    return extracted

@router.get("/memory", response_model=ImportUpdateResponse)
async def get_planner_memory(current_user: UserContext = Depends(get_current_user)):
    try:
        db = get_db()
        imports_cursor = db.planner_memory.find({"user_id": current_user.user_id}).sort("upload_timestamp", -1)
        imports = []
        async for doc in imports_cursor:
            doc["_id"] = str(doc["_id"])
            imports.append(PlannerImport(**doc))
        return {"success": True, "imports": imports}
    except Exception as e:
        logger.error(f"Error in /memory: {e}")
        return {"success": True, "imports": []}

@router.patch("/memory/{import_id}", response_model=ImportUpdateResponse)
async def toggle_import(import_id: str, current_user: UserContext = Depends(get_current_user)):
    global current_schedule, current_insights
    db = get_db()
    doc = await db.planner_memory.find_one({"import_id": import_id, "user_id": current_user.user_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Import not found")
    new_active = not doc.get("active", True)
    await db.planner_memory.update_one({"import_id": import_id, "user_id": current_user.user_id}, {"$set": {"active": new_active}})
    current_schedule = []
    current_insights = []
    return await get_planner_memory(current_user=current_user)

@router.delete("/memory/{import_id}", response_model=ImportUpdateResponse)
async def delete_import(import_id: str, current_user: UserContext = Depends(get_current_user)):
    global current_schedule, current_insights
    db = get_db()
    await db.planner_memory.delete_one({"import_id": import_id, "user_id": current_user.user_id})
    current_schedule = []
    current_insights = []
    return await get_planner_memory(current_user=current_user)

@router.post("/upload")
async def upload_timetable(file: UploadFile = File(...), current_user: UserContext = Depends(get_current_user)):
    """
    Accepts an image, CSV, or XLSX file and extracts tasks.
    Saves it to planner memory.
    """
    try:
        content = await file.read()
        
        extracted_tasks = []
        source_type = "unknown"
        filename_lower = file.filename.lower() if file.filename else ""
        
        if file.content_type.startswith("image/") or filename_lower.endswith(('.png', '.jpg', '.jpeg', '.webp')):
            source_type = "image"
            res = parse_image_timetable(content, mime_type=file.content_type if file.content_type.startswith("image/") else "image/jpeg")
            if not res.get("success"):
                raise HTTPException(status_code=400, detail=res.get("error", "Could not read timetable image clearly"))
            extracted_tasks = res.get("tasks", [])
        elif file.content_type in ["text/csv", "application/vnd.ms-excel"] or filename_lower.endswith('.csv'):
            source_type = "csv"
            extracted_tasks = parse_csv_timetable(content.decode("utf-8", errors="ignore"))
        elif file.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or filename_lower.endswith('.xlsx'):
            source_type = "spreadsheet"
            extracted_tasks = parse_excel_timetable(content)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file.content_type} ({file.filename})")
            
        import_record = PlannerImport(
            file_name=file.filename,
            source_type=source_type,
            parsed_data=extracted_tasks,
            active=True
        )
        import_dict = import_record.dict()
        import_dict["user_id"] = current_user.user_id
        
        try:
            db = get_db()
            # Prevent duplication by filename
            existing = await db.planner_memory.find_one({"file_name": file.filename, "user_id": current_user.user_id})
            if existing:
                await db.planner_memory.update_one(
                    {"file_name": file.filename, "user_id": current_user.user_id}, 
                    {"$set": {"parsed_data": extracted_tasks, "active": True}}
                )
            else:
                await db.planner_memory.insert_one(import_dict)
            
            return await get_planner_memory(current_user=current_user)
        except Exception as db_e:
            logger.error(f"DB Error on upload: {db_e}")
            import_dict["import_id"] = str(uuid.uuid4())
            return {"success": True, "imports": [import_dict]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error parsing upload: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/generate")
async def generate_daily_schedule(req: GenerateRequest, current_user: UserContext = Depends(get_current_user)):
    """
    Takes a mix of existing and newly extracted tasks, 
    scores them, and schedules them.
    """
    global current_schedule, current_insights
    try:
        logger.info(f"[DEBUG] /planner/generate received tasks: {req.tasks}")
        all_tasks = req.tasks.copy()
        
        try:
            db = get_db()
            imports_cursor = db.planner_memory.find({"active": True, "user_id": current_user.user_id})
            async for doc in imports_cursor:
                raw_tasks = doc.get("parsed_data", [])
                for rt in raw_tasks:
                    if isinstance(rt, dict):
                        normalized = {
                            "title": str(rt.get("title", "Untitled")),
                            "category": str(rt.get("category", "Class")),
                            "fixed_or_flexible": str(rt.get("fixed_or_flexible", "Fixed")),
                            "completed": bool(rt.get("completed", False))
                        }
                        if rt.get("day"):
                            normalized["day"] = str(rt.get("day"))
                        if rt.get("start_time"):
                            normalized["start_time"] = str(rt.get("start_time"))
                        if rt.get("end_time"):
                            normalized["end_time"] = str(rt.get("end_time"))
                        if rt.get("duration_minutes"):
                            try:
                                normalized["duration_minutes"] = int(rt.get("duration_minutes"))
                            except:
                                normalized["duration_minutes"] = 60
                        all_tasks.append(normalized)
                        
            calendar_events = await db["calendar_events"].find({"user_id": current_user.user_id}).to_list(100)
            user_doc = await db["users"].find_one({"user_id": current_user.user_id})
            user_prefs = user_doc.get("preferences", {}) if user_doc else {}
        except Exception as db_e:
            logger.warning(f"Skipping DB imports due to error: {db_e}")
            calendar_events = []
            user_prefs = {}

        logger.info(f"[DEBUG] Total tasks before scoring: {len(all_tasks)}")
        
        # 1. Run global batch scheduler which computes & persists the schedule
        from planner.batch_scheduler import schedule_tasks
        await schedule_tasks(current_user.user_id)
        
        # 2. Fetch the newly computed calendar_events
        try:
            calendar_events = await db["calendar_events"].find({"user_id": current_user.user_id}).to_list(100)
        except Exception:
            calendar_events = []
            
        scored = [compute_priority(t) for t in all_tasks]
        schedule = generate_schedule(scored, calendar_events, user_prefs)
        insights = generate_planner_insights(schedule)
        
        current_schedule = schedule
        current_insights = insights
        
        return {
            "success": True,
            "schedule": schedule,
            "insights": insights
        }
    except Exception as e:
        logger.error(f"Error generating schedule: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate schedule")
